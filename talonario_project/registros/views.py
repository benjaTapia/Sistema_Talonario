from django.shortcuts import render, redirect
from django.views.generic import ListView, CreateView, UpdateView, DetailView, DeleteView
from django.urls import reverse_lazy
from .models import Registro, ArchivoAdjunto, Transaccion
from django.contrib import messages
import pandas as pd
from django.http import HttpResponse
from .forms import RegistroForm
from django.urls import reverse
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.views import LoginView
from django.http import Http404
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from django.http import HttpResponseRedirect
from django.core.paginator import Paginator
from django.template.loader import render_to_string
from weasyprint import HTML
from django.db import transaction
from django.contrib.messages import get_messages
from django.db.models import Prefetch
from django.db.models import Sum
from django.utils.formats import number_format
from django.core.paginator import EmptyPage, InvalidPage
from django.db.models import Sum, DecimalField, F, Value
from django.db.models.functions import Coalesce
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import logging
from django.core.exceptions import ValidationError
#================================================================================= VISTAS =======================================================================================
class CustomLoginView(LoginView):
    template_name = 'login.html'

    def dispatch(self, request, *args, **kwargs):
        # Limpiar cualquier mensaje previo
        storage = get_messages(request)
        for _ in storage:
            pass
        return super().dispatch(request, *args, **kwargs)

    def form_invalid(self, form):
        messages.error(self.request, "Usuario o contraseña incorrecta. Por favor, inténtelo de nuevo.")
        return super().form_invalid(form)

    def get_success_url(self):
        messages.success(self.request, "Inicio de sesión correcto. ¡Bienvenido!")
        if self.request.user.is_staff or self.request.user.is_superuser:
            return reverse_lazy('admin:index')
        return reverse_lazy('registros:lista_registros')



#============================================================================= LISTAR REGISTRO ==================================================================================
logger = logging.getLogger(__name__)

class RegistroListView(LoginRequiredMixin, ListView):
    model = Registro
    template_name = 'lista_registros.html'
    context_object_name = 'registros'

    def get_queryset(self):
        """
        Obtiene el conjunto de datos base con filtros aplicados.
        """
        queryset = Registro.objects.filter(usuario=self.request.user).prefetch_related(
            Prefetch(
                'archivos',
                queryset=ArchivoAdjunto.objects.filter(eliminado=False)  # Solo incluir archivos no eliminados
            )
        ).order_by('-fecha_creacion')  # Ordenar por fecha de creación descendente

        # Filtrar por folio
        folio = self.request.GET.get('folio', '').strip()
        if folio:
            queryset = queryset.filter(folio__icontains=folio)

        # Filtrar por nombre
        nombre = self.request.GET.get('nombre', '').strip()
        if nombre:
            queryset = queryset.filter(nombre_registro__icontains=nombre)

        # Filtrar por tipo de registro
        tipo = self.request.GET.get('tipo', '').strip()
        if tipo:
            queryset = queryset.filter(tipo_registro=tipo)

        # Filtrar por rango de fechas
        fecha_inicio = self.request.GET.get('fecha_inicio', '').strip()
        fecha_fin = self.request.GET.get('fecha_fin', '').strip()

        if fecha_inicio and fecha_fin:
            queryset = queryset.filter(fecha_creacion__range=[fecha_inicio, f"{fecha_fin} 23:59:59"])
        elif fecha_inicio:
            queryset = queryset.filter(fecha_creacion__gte=fecha_inicio)
        elif fecha_fin:
            queryset = queryset.filter(fecha_creacion__lte=f"{fecha_fin} 23:59:59")

        return queryset

    def get_context_data(self, **kwargs):
        """
        Divide los registros en 4 tablas: convenio_ingreso, convenio_egreso, aporte_ingreso, aporte_egreso.
        """
        context = super().get_context_data(**kwargs)
        registros = self.get_queryset()

        # Log para depuración
        for registro in registros:
            logger.info(f"Registro: {registro.id}, Archivos válidos: {[archivo.archivo.name for archivo in registro.archivos.all()]}")

        # Dividir registros según el tipo
        convenio_ingreso = registros.filter(tipo_registro='convenio_ingreso')
        convenio_egreso = registros.filter(tipo_registro='convenio_egreso')
        aporte_ingreso = registros.filter(tipo_registro='aporte_ingreso')
        aporte_egreso = registros.filter(tipo_registro='aporte_egreso')

        # Paginación para cada tipo de registro
        paginator_convenio_ingreso = Paginator(convenio_ingreso, 5)
        paginator_convenio_egreso = Paginator(convenio_egreso, 5)
        paginator_aporte_ingreso = Paginator(aporte_ingreso, 5)
        paginator_aporte_egreso = Paginator(aporte_egreso, 5)

        # Obtener la página actual para cada tabla
        page_convenio_ingreso = self.request.GET.get('page_convenio_ingreso', 1)
        page_convenio_egreso = self.request.GET.get('page_convenio_egreso', 1)
        page_aporte_ingreso = self.request.GET.get('page_aporte_ingreso', 1)
        page_aporte_egreso = self.request.GET.get('page_aporte_egreso', 1)

        # Manejar excepciones de paginación para evitar errores
        try:
            context['convenio_ingreso_paginated'] = paginator_convenio_ingreso.get_page(page_convenio_ingreso)
        except (EmptyPage, InvalidPage):
            context['convenio_ingreso_paginated'] = paginator_convenio_ingreso.get_page(1)

        try:
            context['convenio_egreso_paginated'] = paginator_convenio_egreso.get_page(page_convenio_egreso)
        except (EmptyPage, InvalidPage):
            context['convenio_egreso_paginated'] = paginator_convenio_egreso.get_page(1)

        try:
            context['aporte_ingreso_paginated'] = paginator_aporte_ingreso.get_page(page_aporte_ingreso)
        except (EmptyPage, InvalidPage):
            context['aporte_ingreso_paginated'] = paginator_aporte_ingreso.get_page(1)

        try:
            context['aporte_egreso_paginated'] = paginator_aporte_egreso.get_page(page_aporte_egreso)
        except (EmptyPage, InvalidPage):
            context['aporte_egreso_paginated'] = paginator_aporte_egreso.get_page(1)

        return context



#============================================================================= CREAR REGISTRO ===================================================================================
class RegistroCreateView(CreateView):
    model = Registro
    template_name = 'crear_registro.html'
    form_class = RegistroForm  # Utilizamos el formulario personalizado
    success_url = reverse_lazy('registros:lista_registros')

    def form_valid(self, form):
        try:
            with transaction.atomic():
                # Asignar el usuario actual al registro
                registro = form.save(commit=False)
                form.instance.usuario = self.request.user  

                # Procesar valores de "Debe" y "Haber"
                debe = self.request.POST.getlist('debe[]')
                haber = self.request.POST.getlist('haber[]')

                # Validar y calcular los totales de "Debe" y "Haber"
                total_debe = sum(float(value.replace(",", "")) for value in debe if value.strip())
                total_haber = sum(float(value.replace(",", "")) for value in haber if value.strip())

                if total_debe + total_haber <= 0:
                    form.add_error(None, "Debe ingresar al menos un valor positivo en 'Debe' o 'Haber'.")
                    return self.form_invalid(form)

                # Asignar el monto como suma de "Debe" y "Haber"
                registro.monto = total_debe + total_haber

                # Guardar el registro
                registro.debe = total_debe
                registro.haber = total_haber
                registro.save()

                # Guardar transacciones en la base de datos
                transacciones = [
                    Transaccion(registro=registro, debe=float(value), haber=0) for value in debe if value.strip()
                ] + [
                    Transaccion(registro=registro, debe=0, haber=float(value)) for value in haber if value.strip()
                ]
                Transaccion.objects.bulk_create(transacciones)

                # Procesar archivos adjuntos
                archivos = self.request.FILES.getlist('archivos[]')
                descripciones = self.request.POST.getlist('descripciones[]')
                if len(descripciones) < len(archivos):
                    descripciones.extend(["Sin descripción"] * (len(archivos) - len(descripciones)))

                adjuntos = [
                    ArchivoAdjunto(
                        registro=registro,
                        archivo=archivo,
                        descripcion=descripcion.strip() if descripcion.strip() else "Sin descripción"
                    )
                    for archivo, descripcion in zip(archivos, descripciones)
                ]
                ArchivoAdjunto.objects.bulk_create(adjuntos)

                messages.success(self.request, "Registro creado exitosamente.")
                return HttpResponseRedirect(reverse('registros:detalle_registro', kwargs={'pk': registro.pk}))

        except Exception as e:
            form.add_error(None, f"Error inesperado: {str(e)}")
            return self.form_invalid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Hubo un error al crear el registro. Por favor revisa los campos.")
        return super().form_invalid(form)



#============================================================================= EDITAR REGISTRO ==================================================================================
class RegistroUpdateView(UpdateView):
    model = Registro
    template_name = 'editar_registro.html'
    form_class = RegistroForm  # Utilizamos el formulario personalizado

    def form_valid(self, form):
        try:
            with transaction.atomic():
                registro = form.save(commit=False)

                # Restaurar el folio original (evita la modificación del folio)
                registro.folio = self.object.folio

                # Procesar los valores de "debe" y "haber" desde el formulario
                debe = self.request.POST.getlist('debe[]')
                haber = self.request.POST.getlist('haber[]')

                # Validar y calcular los totales de "debe" y "haber"
                try:
                    total_debe = sum(float(value.replace(',', '')) for value in debe if value.strip())
                    total_haber = sum(float(value.replace(',', '')) for value in haber if value.strip())
                except ValueError:
                    form.add_error(None, "Los valores de 'Debe' y 'Haber' deben ser numéricos.")
                    return self.form_invalid(form)

                # Asignar los totales al registro
                registro.debe = total_debe
                registro.haber = total_haber
                registro.save()

                # Actualizar las transacciones existentes
                registro.transacciones.all().delete()
                transacciones = [
                    Transaccion(registro=registro, debe=float(value.replace(',', '').strip()), haber=0.0)
                    for value in debe if value.strip()
                ] + [
                    Transaccion(registro=registro, debe=0.0, haber=float(value.replace(',', '').strip()))
                    for value in haber if value.strip()
                ]
                Transaccion.objects.bulk_create(transacciones)

                # Procesar eliminación de archivos existentes
                eliminar_adjuntos = self.request.POST.getlist('eliminar_adjuntos')
                if eliminar_adjuntos:
                    for archivo_id in eliminar_adjuntos:
                        archivo = ArchivoAdjunto.objects.get(id=archivo_id, registro=registro)
                        archivo.eliminar_archivo()  # Llamar a la función eliminar_archivo

                # Guardar nuevos archivos adjuntos
                nuevos_archivos = self.request.FILES.getlist('archivos[]')
                nuevas_descripciones = self.request.POST.getlist('descripciones[]')
                if nuevos_archivos:
                    adjuntos = [
                        ArchivoAdjunto(
                            registro=registro,
                            archivo=archivo,
                            descripcion=descripcion.strip() if descripcion.strip() else "Sin descripción"
                        )
                        for archivo, descripcion in zip(nuevos_archivos, nuevas_descripciones)
                    ]
                    ArchivoAdjunto.objects.bulk_create(adjuntos)

            messages.success(self.request, "Registro actualizado exitosamente.")
            return super().form_valid(form)

        except Exception as e:
            form.add_error(None, f"Ha ocurrido un error: {str(e)}")
            return self.form_invalid(form)

    def get_success_url(self):
        """
        Redirige al detalle del registro actualizado después de guardar los cambios.
        """
        return reverse('registros:detalle_registro', args=[self.object.pk])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['transacciones_debe'] = self.object.transacciones.filter(debe__gt=0)
        context['transacciones_haber'] = self.object.transacciones.filter(haber__gt=0)
        context['archivos'] = self.object.archivos.filter(eliminado=False)  # Solo mostrar archivos activos
        return context



#============================================================================ DETALLE REGISTRO ==================================================================================
class RegistroDetailView(LoginRequiredMixin, DetailView):
    model = Registro
    template_name = 'detalle_registro.html'
    context_object_name = 'registro'

    def get_queryset(self):
        """
        Filtra los registros para mostrar solo aquellos pertenecientes al usuario autenticado.
        """
        return Registro.objects.filter(usuario=self.request.user)

    def get_object(self, queryset=None):
        """
        Obtiene el objeto y verifica que pertenezca al usuario autenticado.
        """
        obj = super().get_object(queryset)
        if obj.usuario != self.request.user:
            raise Http404("No tienes permiso para ver este registro.")
        return obj

    def get_context_data(self, **kwargs):
        """
        Añade información adicional al contexto, incluyendo los parámetros para volver a la lista
        en la pestaña y página correcta.
        """
        context = super().get_context_data(**kwargs)
        
        # Filtrar solo archivos no eliminados
        context['archivos'] = ArchivoAdjunto.objects.filter(
            registro=self.object, eliminado=False
        )
        
        # Calcular totales de debe y haber
        total_debe = sum(transaccion.debe for transaccion in self.object.transacciones.all())
        total_haber = sum(transaccion.haber for transaccion in self.object.transacciones.all())
        context['total_debe'] = total_debe
        context['total_haber'] = total_haber

        # Agregar parámetros para volver a la lista
        context['active_tab'] = self.request.GET.get('active_tab', 'convenio-ingreso')  # Pestaña activa por defecto
        context['page'] = self.request.GET.get('page', 1)  # Página activa por defecto

        return context



#============================================================================ ELIMINAR REGISTRO =================================================================================
class RegistroDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Registro
    template_name = 'confirmar_eliminacion.html'
    success_url = reverse_lazy('registros:lista_registros')
    permission_required = 'registros.delete_registro'

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Registro eliminado exitosamente.")
        return super().delete(request, *args, **kwargs)



#=========================================================================== HISTORIAL ARCHIVOS =================================================================================
class HistorialArchivosEliminadosView(ListView):
    model = ArchivoAdjunto
    template_name = 'historial_archivos.html'  # Template que crearemos
    context_object_name = 'archivos_eliminados'

    def get_queryset(self):
        return ArchivoAdjunto.objects.filter(eliminado=True).order_by('-fecha_eliminacion')



#============================================================================= REPORTES EXCEL ===================================================================================
def exportar_excel(request):
    registros = Registro.objects.all().values(
        'folio', 'nombre_registro', 'tipo_registro', 'fecha_creacion', 'detalle', 'monto'
    )
    df = pd.DataFrame(registros)

    # Reformatear la columna 'folio' para separar por tipo de registro
    if 'folio' in df.columns and 'tipo_registro' in df.columns:
        df['folio'] = df.apply(
            lambda row: f"{row['tipo_registro'].split('_')[0].capitalize()} - {row['folio']}", axis=1
        )

    # Convertir fechas al formato adecuado
    if 'fecha_creacion' in df.columns:
        df['fecha_creacion'] = pd.to_datetime(df['fecha_creacion']).dt.tz_localize(None)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="registros.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Registros')

        worksheet = writer.sheets['Registros']

        # Estilos para encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Aplicar estilos a los encabezados
        for cell in worksheet[1]:  # Primera fila contiene los encabezados
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Ajustar ancho de columnas de forma responsiva
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Obtener la letra de la columna
            for cell in column:
                try:
                    if cell.value:
                        value_length = len(str(cell.value))
                        # Ajustar el ancho para contenido más largo
                        max_length = max(max_length, value_length)
                except:
                    pass
            # Ajustar ancho con un factor dinámico para mejorar la visibilidad
            worksheet.column_dimensions[column_letter].width = max_length + 3  # Ajuste dinámico

        # Aplicar formato numérico a la columna 'monto' si existe
        if 'monto' in df.columns:
            monto_col_idx = list(df.columns).index('monto') + 1
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=monto_col_idx)
                if cell.value is not None:
                    try:
                        cell.value = float(cell.value)
                        cell.number_format = '#,##0'  # Formato con separadores de miles
                    except ValueError:
                        pass

    return response



#============================================================================= EDITAR PERFIL ====================================================================================
@login_required
def editar_perfil(request):
    user = request.user

    # Diccionario para almacenar errores de validación
    errors = {}

    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()

        # Validaciones específicas
        if not email:
            errors['email'] = "El campo 'Correo Electrónico' es obligatorio."
        if not first_name:
            errors['first_name'] = "El campo 'Nombre' es obligatorio."
        if not last_name:
            errors['last_name'] = "El campo 'Apellido' es obligatorio."

        # Si hay errores, renderizar el formulario con mensajes de error
        if errors:
            return render(request, 'editar_perfil.html', {
                'user': user,
                'errors': errors
            })

        try:
            # Actualiza los datos del usuario
            user.email = email
            user.first_name = first_name
            user.last_name = last_name
            user.full_clean()  # Valida el modelo antes de guardar
            user.save()

            messages.success(request, "Perfil actualizado exitosamente.")
            return redirect('registros:lista_registros')

        except ValidationError as e:
            # Capturar errores de validación y mostrarlos en el formulario
            for field, error_list in e.message_dict.items():
                errors[field] = ', '.join(error_list)
            return render(request, 'editar_perfil.html', {
                'user': user,
                'errors': errors
            })

        except Exception as e:
            # Manejo general de errores
            messages.error(request, f"Error inesperado: {str(e)}")
            return render(request, 'editar_perfil.html', {'user': user, 'errors': errors})

    # Renderizar formulario inicial sin errores
    return render(request, 'editar_perfil.html', {'user': user, 'errors': {}})



#======================================================================== ADJUNTAR DOCUMENTOS FIRMADOS ==========================================================================
def adjuntar_firmado(request, pk):
    registro = get_object_or_404(Registro, pk=pk)

    if request.method == 'POST' and request.FILES.get('archivo_firmado'):
        archivo_firmado = request.FILES['archivo_firmado']
        ArchivoAdjunto.objects.create(
            registro=registro,
            archivo=archivo_firmado,
            descripcion="Documento firmado"
        )
        return HttpResponseRedirect(reverse('registros:detalle_registro', args=[pk]))

    return render(request, 'adjuntar_firmado.html', {'registro': registro})



#================================================================================ GENERAR PDF ===================================================================================
def generar_pdf(request, pk):
    # Obtén el registro
    registro = get_object_or_404(Registro, pk=pk)

    # Calcular los totales basado en los campos 'debe' y 'haber'
    total_debe = registro.transacciones.aggregate(total=Sum('debe'))['total'] or 0.0
    total_haber = registro.transacciones.aggregate(total=Sum('haber'))['total'] or 0.0

    # Formatear los números al estilo chileno
    registro.monto_formateado = number_format(registro.monto, decimal_pos=2, use_l10n=True)
    total_debe_formateado = number_format(total_debe, decimal_pos=0, use_l10n=True)
    total_haber_formateado = number_format(total_haber, decimal_pos=0, use_l10n=True)

    # Pasar los valores formateados al contexto
    context = {
        'registro': registro,
        'total_debe': total_debe_formateado,
        'total_haber': total_haber_formateado,
    }

    # Renderizar el PDF
    html_string = render_to_string('pdf/detalle_registro_pdf.html', context)
    pdf = HTML(string=html_string).write_pdf()

    # Retornar el PDF como respuesta
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename=\"registro_{registro.folio}.pdf\"'
    return response



#============================================================================== LIBRO CONTABLE ==================================================================================
@login_required
def contabilidad_view(request):
    # Filtrar registros por usuario actual
    usuario = request.user

    convenios = Registro.objects.filter(
        usuario=usuario,  # Filtrar registros del usuario actual
        tipo_registro__startswith='convenio'
    ).annotate(
        debe=Coalesce(Sum('transacciones__debe'), Value(0, output_field=DecimalField())),
        haber=Coalesce(Sum('transacciones__haber'), Value(0, output_field=DecimalField())),
        saldo=Coalesce(Sum(F('transacciones__debe') - F('transacciones__haber')), Value(0, output_field=DecimalField()))
    )

    aportes = Registro.objects.filter(
        usuario=usuario,  # Filtrar registros del usuario actual
        tipo_registro__startswith='aporte'
    ).annotate(
        debe=Coalesce(Sum('transacciones__debe'), Value(0, output_field=DecimalField())),
        haber=Coalesce(Sum('transacciones__haber'), Value(0, output_field=DecimalField())),
        saldo=Coalesce(Sum(F('transacciones__debe') - F('transacciones__haber')), Value(0, output_field=DecimalField()))
    )

    # Configuración del paginador (5 registros por página)
    convenios_paginator = Paginator(convenios, 5)
    aportes_paginator = Paginator(aportes, 5)

    # Obtener la página actual de convenios y aportes
    convenios_page_number = request.GET.get('convenios_page')
    aportes_page_number = request.GET.get('aportes_page')

    convenios_page = convenios_paginator.get_page(convenios_page_number)
    aportes_page = aportes_paginator.get_page(aportes_page_number)

    # Determinar la pestaña activa
    active_tab = request.GET.get('active_tab', 'convenios')

    # Calcular el Estado de Resultados
    ingresos = Registro.objects.filter(
        usuario=usuario,  # Filtrar por usuario
        tipo_registro__icontains='ingreso'
    ).aggregate(
        total=Coalesce(Sum('monto'), Value(0, output_field=DecimalField()))
    )['total']

    egresos = Registro.objects.filter(
        usuario=usuario,  # Filtrar por usuario
        tipo_registro__icontains='egreso'
    ).aggregate(
        total=Coalesce(Sum('monto'), Value(0, output_field=DecimalField()))
    )['total']

    ganancia_perdida = ingresos - egresos

    estado_resultados = {
        'ingresos': ingresos,
        'egresos': egresos,
        'ganancia_perdida': ganancia_perdida,
    }

    # Contexto para la plantilla
    context = {
        'libro_mayor_convenios': convenios_page,
        'libro_mayor_aportes': aportes_page,
        'estado_resultados': estado_resultados,
        'active_tab': active_tab,  # Indicar la pestaña activa
    }

    return render(request, 'contabilidad.html', context)



#========================================================================== EXPORTAR EXCEL CONTABLE =============================================================================
def exportar_excel_contabilidad(request):
    # Crear el libro y la hoja de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contabilidad"

    # Estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Agregar encabezados
    headers = ["Fecha", "Folio", "Descripción", "Debe", "Haber", "Saldo"]
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Obtener datos de contabilidad
    convenios = Registro.objects.filter(tipo_registro__startswith='convenio').annotate(
        debe=Coalesce(Sum('transacciones__debe'), Value(0, output_field=DecimalField())),
        haber=Coalesce(Sum('transacciones__haber'), Value(0, output_field=DecimalField())),
        saldo=Coalesce(Sum(F('transacciones__debe') - F('transacciones__haber')), Value(0, output_field=DecimalField()))
    )

    # Agregar filas con datos
    row_num = 2
    for registro in convenios:
        ws.cell(row=row_num, column=1, value=registro.fecha_creacion.strftime("%Y-%m-%d"))
        ws.cell(row=row_num, column=2, value=registro.folio)
        ws.cell(row=row_num, column=3, value=registro.nombre_registro)
        ws.cell(row=row_num, column=4, value=float(registro.debe))
        ws.cell(row=row_num, column=5, value=float(registro.haber))
        ws.cell(row=row_num, column=6, value=float(registro.saldo))

        # Aplicar bordes y alineación a las celdas de datos
        for col_num in range(1, 7):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Aplicar formato de número con separadores de miles
        ws.cell(row=row_num, column=4).number_format = '#,##0'  # Formato sin decimales
        ws.cell(row=row_num, column=5).number_format = '#,##0'
        ws.cell(row=row_num, column=6).number_format = '#,##0'

        # Colorear valores negativos en rojo
        if registro.saldo < 0:
            ws.cell(row=row_num, column=6).font = Font(color="FF0000")

        row_num += 1

    # Ajustar el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except TypeError:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Configurar la respuesta HTTP
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="contabilidad.xlsx"'
    wb.save(response)

    return response



#=========================================================================== EXPORTAR PDF CONTABLE ==============================================================================
def exportar_pdf_contabilidad(request):
    # Filtrar registros para Convenios Ingreso
    convenios_ingreso = Registro.objects.filter(tipo_registro='convenio_ingreso').annotate(
        debe=Coalesce(Sum('transacciones__debe'), Value(0, output_field=DecimalField())),
        haber=Coalesce(Sum('transacciones__haber'), Value(0, output_field=DecimalField())),
        saldo=Coalesce(Sum(F('transacciones__debe') - F('transacciones__haber')), Value(0, output_field=DecimalField()))
    )

    # Filtrar registros para Convenios Egreso
    convenios_egreso = Registro.objects.filter(tipo_registro='convenio_egreso').annotate(
        debe=Coalesce(Sum('transacciones__debe'), Value(0, output_field=DecimalField())),
        haber=Coalesce(Sum('transacciones__haber'), Value(0, output_field=DecimalField())),
        saldo=Coalesce(Sum(F('transacciones__debe') - F('transacciones__haber')), Value(0, output_field=DecimalField()))
    )

    # Filtrar registros para Aportes Ingreso
    aportes_ingreso = Registro.objects.filter(tipo_registro='aporte_ingreso').annotate(
        debe=Coalesce(Sum('transacciones__debe'), Value(0, output_field=DecimalField())),
        haber=Coalesce(Sum('transacciones__haber'), Value(0, output_field=DecimalField())),
        saldo=Coalesce(Sum(F('transacciones__debe') - F('transacciones__haber')), Value(0, output_field=DecimalField()))
    )

    # Filtrar registros para Aportes Egreso
    aportes_egreso = Registro.objects.filter(tipo_registro='aporte_egreso').annotate(
        debe=Coalesce(Sum('transacciones__debe'), Value(0, output_field=DecimalField())),
        haber=Coalesce(Sum('transacciones__haber'), Value(0, output_field=DecimalField())),
        saldo=Coalesce(Sum(F('transacciones__debe') - F('transacciones__haber')), Value(0, output_field=DecimalField()))
    )

    # Calcular Estado de Resultados
    ingresos = Registro.objects.filter(tipo_registro__icontains='ingreso').aggregate(
        total=Coalesce(Sum('monto'), Value(0, output_field=DecimalField()))
    )['total']

    egresos = Registro.objects.filter(tipo_registro__icontains='egreso').aggregate(
        total=Coalesce(Sum('monto'), Value(0, output_field=DecimalField()))
    )['total']

    estado_resultados = {
        'ingresos': ingresos,
        'egresos': egresos,
        'ganancia_perdida': ingresos - egresos,
    }

    # Renderizar plantilla para el PDF
    html_string = render_to_string('pdf/contabilidad_pdf.html', {
        'convenios_ingreso': convenios_ingreso,
        'convenios_egreso': convenios_egreso,
        'aportes_ingreso': aportes_ingreso,
        'aportes_egreso': aportes_egreso,
        'estado_resultados': estado_resultados,
    })

    # Generar PDF con WeasyPrint
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="contabilidad.pdf"'
    HTML(string=html_string).write_pdf(response)

    return response



#============================================================================== PAGINAS DE ERRORES ==============================================================================

def error_404_view(request, exception):
    return render(request, '404.html', status=404)

def error_500_view(request):
    return render(request, '500.html', status=500)
